"""
saa_etl.py
==========
Extract-Transform-Load layer for the SAA integrated assessment (Question 4).

Reads the annual-financial-statement workbook (SAA_Financial_statement.xlsx),
which stores each statement as a separate sheet in published-report layout
(one row per line item, one column per entity-year), and flattens it into a
tidy long-format dataset:

    statement | entity | year | line_item | metric | value_rm

Every figure is in R million, as published. Where FY2017 appears in two
sheets (the 2015-2017 sheet and the restated 2017-2019 sheet), the later
restated figure wins and the difference is written to a separate
restatements file, because restatement deltas are themselves forensic
evidence (Question 2.2 - reliability of the financial statements).

Usage
-----
    python saa_etl.py --source SAA_Financial_statement.xlsx --outdir data
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# 1. Sheet map: which spreadsheet column holds which entity and which year.
#    Column numbers are 1-based (openpyxl convention). Priority breaks ties
#    for years that appear in more than one sheet - higher wins.
# --------------------------------------------------------------------------
SHEET_SPECS: list[dict] = [
    {
        "sheet": "Income statement",
        "statement": "Income statement",
        "priority": 3,
        "columns": {
            3: ("Group", 2019), 4: ("Group", 2018), 5: ("Group", 2017),
            6: ("Company", 2019), 7: ("Company", 2018), 8: ("Company", 2017),
        },
    },
    {
        "sheet": "Income statement 2015-2017",
        "statement": "Income statement",
        "priority": 2,
        "columns": {
            3: ("Group", 2017), 4: ("Group", 2016), 5: ("Group", 2015),
            6: ("Company", 2017), 7: ("Company", 2016), 8: ("Company", 2015),
        },
    },
    {
        "sheet": "income state 2011-2012",
        "statement": "Income statement",
        "priority": 1,
        "columns": {
            3: ("Group", 2012), 4: ("Group", 2011),
            5: ("Company", 2012), 6: ("Company", 2011),
        },
    },
    {
        "sheet": "Balance Sheet",
        "statement": "Balance sheet",
        "priority": 3,
        "columns": {
            3: ("Group", 2019), 4: ("Group", 2018), 5: ("Group", 2017),
            6: ("Company", 2019), 7: ("Company", 2018), 8: ("Company", 2017),
        },
    },
    {
        "sheet": "Balance sheet 2015-2017",
        "statement": "Balance sheet",
        "priority": 2,
        "columns": {
            3: ("Group", 2017), 4: ("Group", 2016), 5: ("Group", 2015),
            6: ("Company", 2017), 7: ("Company", 2016), 8: ("Company", 2015),
        },
    },
    {
        "sheet": "balance sheet 2010-2012",
        "statement": "Balance sheet",
        "priority": 1,
        "columns": {
            3: ("Group", 2012), 4: ("Group", 2011), 5: ("Group", 2010),
            6: ("Company", 2012), 7: ("Company", 2011), 8: ("Company", 2010),
        },
    },
    {
        "sheet": "cash flow",
        "statement": "Cash flow",
        "priority": 3,
        "columns": {
            3: ("Group", 2019), 4: ("Group", 2018), 5: ("Group", 2017),
            6: ("Company", 2019), 7: ("Company", 2018), 8: ("Company", 2017),
        },
    },
    {
        "sheet": "cash flow2015-2017",
        "statement": "Cash flow",
        "priority": 2,
        "columns": {
            3: ("Group", 2017), 4: ("Group", 2016), 5: ("Group", 2015),
            6: ("Company", 2017), 7: ("Company", 2016), 8: ("Company", 2015),
        },
    },
    {
        "sheet": "cash flow2011-2012",
        "statement": "Cash flow",
        "priority": 1,
        "columns": {
            3: ("Group", 2012), 4: ("Group", 2011),
            5: ("Company", 2012), 6: ("Company", 2011),
        },
    },
]

# Balance-sheet section headers. Unlabelled numeric rows are the section
# subtotal in published layout, so the parser attributes them to the section
# that is currently open.
SECTION_HEADERS = {
    "non-current assets": "total non-current assets",
    "current assets": "total current assets",
    "equity attributable to equity holders of parent": "total equity",
    "equity attributable to equity holders of the parent": "total equity",
    "equity attributable to equity holders": "total equity",
    "non-current liabilities": "total non-current liabilities",
    "liabilities non-current liabilities": "total non-current liabilities",
    "current liabilities": "total current liabilities",
}

# Labels that repeat in more than one balance-sheet section and therefore
# need the section name attached to stay unambiguous.
# Structural/header text that must never be glued onto the next line item.
NOISE_LABELS = {
    "r million", "column1", "notes", "assets", "equity", "liabilities",
    "equity and liabilities", "other comprehensive income:",
    "other comprehensive income/(loss):",
    "cash flows from operating activities",
    "cash flows from investing activities",
    "cash flows from financing activities",
    "total comprehensive loss attributable to:",
    "total comprehensive income/(loss) attributable to:",
    "equity attributable to equity holders",
}

AMBIGUOUS_LABELS = {
    "provisions",
    "deferred revenue on ticket sales",
    "aircraft and other deposits",
    "amounts receivable from subsidiaries",
    "investments in subsidiaries",
    "derivatives",
}

# --------------------------------------------------------------------------
# 2. Canonical metric map. Published wording changed over the decade
#    (2012 says "Revenue"/"Turnover"; 2019 says "Total income"/"Airline
#    revenue"), so aliases are folded onto one machine-readable key.
# --------------------------------------------------------------------------
METRIC_ALIASES: dict[str, str] = {
    # ---- Income statement -------------------------------------------------
    "total income": "total_income",
    "revenue": "total_income",
    "airline revenue": "airline_revenue",
    "turnover": "airline_revenue",
    "other income": "other_income",
    "operating costs": "operating_costs",
    "aircraft lease costs": "lease_costs",
    "fuel and other energy costs": "fuel_costs",
    "energy": "fuel_costs",
    "employee benefit expenses": "employee_costs",
    "maintenance costs": "maintenance_costs",
    "material": "maintenance_costs",
    "accommodation and refreshments": "catering_costs",
    "navigation, landing and parking fees": "navigation_costs",
    "commissions and network charges": "commission_costs",
    "distribution costs": "commission_costs",
    "other operating costs": "other_operating_costs",
    "operating loss before interest, tax, depreciation and amortisation": "ebitda",
    "operating (loss)/profit before interest, tax, depreciation and amortisation": "ebitda",
    "operating (loss)/profit before fair value movements and translation profit/(loss)": "ebitda",
    "depreciation and amortisation": "depreciation",
    "impairments": "impairments",
    "net impairment write-off": "impairments",
    "operating loss": "operating_loss",
    "(loss)/profit before finance costs and investment income": "operating_loss",
    "finance costs": "finance_costs",
    "interest income": "interest_income",
    "investment income": "interest_income",
    "loss before taxation": "loss_before_tax",
    "(loss)/profit before taxation": "loss_before_tax",
    "taxation": "taxation",
    "loss for the year": "loss_for_year",
    "(loss)/profit for the year": "loss_for_year",
    "total comprehensive loss": "total_comprehensive_loss",
    "total comprehensive income/(loss)": "total_comprehensive_loss",
    # ---- Balance sheet ----------------------------------------------------
    "total assets": "total_assets",
    "total liabilities": "total_liabilities",
    "total equity and liabilities": "total_equity_and_liabilities",
    "total equity": "total_equity",
    "total non-current assets": "non_current_assets",
    "total current assets": "current_assets",
    "total non-current liabilities": "non_current_liabilities",
    "total current liabilities": "current_liabilities",
    "property, aircraft and equipment": "ppe",
    "intangible assets": "intangibles",
    "inventories": "inventories",
    "trade and other receivables": "receivables",
    "cash and cash equivalents": "cash",
    "share capital": "share_capital",
    "shareholder contribution": "shareholder_contribution",
    "reserves": "reserves",
    "accumulated loss": "accumulated_loss",
    "long-term loans": "long_term_loans",
    "current portion of long-term loans": "current_portion_loans",
    "trade and other payables": "payables",
    "bank overdraft": "bank_overdraft",
    "current liabilities: provisions": "provisions_current",
    "non-current liabilities: provisions": "provisions_non_current",
    "current liabilities: deferred revenue on ticket sales": "deferred_revenue_current",
    "non-current liabilities: deferred revenue on ticket sales": "deferred_revenue_non_current",
    "subordinated loan guaranteed by government": "subordinated_loan",
    # ---- Cash flow --------------------------------------------------------
    "cash (used in)/generated from operations": "cash_from_operations",
    "cash generated from/(used in) operations": "cash_from_operations",
    "net cash outflow from operating activities": "net_operating_cash",
    "net cash (outflow)/inflow from operating activities": "net_operating_cash",
    "net cash inflow from operating activities": "net_operating_cash",
    "net cash outflow from investing activities": "net_investing_cash",
    "net cash inflow from investing activities": "net_investing_cash",
    "net cash inflow from financing activities": "net_financing_cash",
    "net cash outflow from financing activities": "net_financing_cash",
    "additions to property, aircraft and equipment": "capex",
    "additions to intangible assets": "intangible_additions",
    "external borrowings raised": "borrowings_raised",
    "external borrowings repaid": "borrowings_repaid",
    "movement in bank overdraft": "overdraft_movement",
    "proceeds from contribution made by the shareholder during the year": "shareholder_bailout",
    "tax paid": "tax_paid",
    "cash and cash equivalents at the end of the year": "closing_cash",
    "cash and cash equivalents at end of the year": "closing_cash",
}

# Metrics published as negatives in some years and positives in others
# (costs deducted vs costs listed). Stored as magnitudes so that trends are
# comparable across the decade.
ABSOLUTE_METRICS = {
    "depreciation", "finance_costs", "capex", "operating_costs",
    "lease_costs", "fuel_costs", "employee_costs", "maintenance_costs",
    "catering_costs", "navigation_costs", "commission_costs",
    "other_operating_costs", "tax_paid",
}


# --------------------------------------------------------------------------
# 3. Cell-level parsing helpers
# --------------------------------------------------------------------------
def normalise(text: str) -> str:
    """Fold published label variants onto one comparable string."""
    if text is None:
        return ""
    text = unicodedata.normalize("NFKC", str(text))
    text = text.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
    text = text.replace("\xa0", " ").replace("\n", " ")
    text = re.sub(r"[*†‡]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def parse_value(cell) -> float | None:
    """Convert a published cell to a float.

    Handles the four conventions in this workbook: real numbers, thin-space
    thousands separators ("26 992"), accounting negatives ("(558)"), and the
    en-dash used for nil.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)

    raw = unicodedata.normalize("NFKC", str(cell)).strip()
    if raw in {"", "-", "\u2013", "\u2014", "n/a", "na"}:
        return None
    # A header string that leaked into a data column.
    if raw.lower().startswith("column"):
        return None

    negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()")
    raw = raw.replace("\xa0", "").replace(" ", "").replace(",", "")
    raw = raw.replace("\u2013", "-").replace("\u2212", "-")
    try:
        value = float(raw)
    except ValueError:
        return None
    return -value if negative else value


# --------------------------------------------------------------------------
# 4. Sheet reader
# --------------------------------------------------------------------------
def read_sheet(workbook, spec: dict) -> list[dict]:
    """Flatten one published statement sheet into tidy records."""
    ws = workbook[spec["sheet"]]
    records: list[dict] = []
    current_section: str | None = None
    pending_label = ""

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label_raw = row[0] if len(row) else None
        label = normalise(label_raw)
        values = {
            col: parse_value(row[col - 1]) if len(row) >= col else None
            for col in spec["columns"]
        }
        has_values = any(v is not None for v in values.values())

        # Section header (balance sheet only): remember it, emit nothing.
        if spec["statement"] == "Balance sheet" and label in SECTION_HEADERS:
            current_section = label
            pending_label = ""
            continue

        # Label wrapped onto its own row, e.g. "Cash and cash equivalents at
        # the beginning of" / "the year". Carry it to the next row.
        if label and not has_values:
            pending_label = "" if label in NOISE_LABELS else label
            continue

        if not has_values:
            continue

        # A label may have wrapped onto the previous row. Prefer whichever
        # form is a recognised line item; fall back to the merged text.
        merged = f"{pending_label} {label}".strip() if pending_label else label
        if merged in METRIC_ALIASES:
            label = merged
        elif label in METRIC_ALIASES or pending_label in NOISE_LABELS:
            pass
        else:
            label = merged
        pending_label = ""

        # Unlabelled numeric row inside a balance-sheet section = subtotal.
        if not label:
            if spec["statement"] == "Balance sheet" and current_section:
                label = SECTION_HEADERS[current_section]
            else:
                continue
        elif label in AMBIGUOUS_LABELS and current_section:
            label = f"{current_section}: {label}"

        metric = METRIC_ALIASES.get(label)
        for col, (entity, year) in spec["columns"].items():
            value = values.get(col)
            if value is None:
                continue
            if metric in ABSOLUTE_METRICS:
                value = abs(value)
            records.append({
                "statement": spec["statement"],
                "entity": entity,
                "year": year,
                "line_item": label_raw if label_raw else label,
                "label_key": label,
                "metric": metric,
                "value_rm": value,
                "source_sheet": spec["sheet"],
                "priority": spec["priority"],
            })
    return records


def build_dataset(source: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (tidy dataset, restatement differences)."""
    wb = load_workbook(source, read_only=True, data_only=True)
    rows: list[dict] = []
    for spec in SHEET_SPECS:
        if spec["sheet"] not in wb.sheetnames:
            print(f"  ! sheet not found, skipped: {spec['sheet']}")
            continue
        found = read_sheet(wb, spec)
        print(f"  - {spec['sheet']:<28} {len(found):>4} values")
        rows.extend(found)
    wb.close()

    df = pd.DataFrame(rows)

    # FY2017 is published twice. Keep the later (restated) figure, but record
    # what changed - a restatement of this size is itself a red flag.
    # Dedupe on the canonical metric where one exists, so that a line item
    # renamed between reports (e.g. "Operating loss before interest, tax,
    # depreciation and amortisation" vs "Operating (loss)/profit before ...")
    # is still recognised as the same measure.
    df["dedup_key"] = df["metric"].fillna(df["label_key"])
    key = ["statement", "entity", "year", "dedup_key"]
    duplicated = df[df.duplicated(key, keep=False)].copy()
    restatements = pd.DataFrame()
    if not duplicated.empty:
        pivot = (duplicated.sort_values("priority", kind="stable")
                 .groupby(key)
                 .agg(original_rm=("value_rm", "first"),
                      restated_rm=("value_rm", "last"),
                      original_sheet=("source_sheet", "first"),
                      restated_sheet=("source_sheet", "last"))
                 .reset_index())
        pivot["difference_rm"] = pivot["restated_rm"] - pivot["original_rm"]
        # Same-sheet duplicates (a section subtotal repeated in the published
        # layout) are not restatements - only cross-report differences are.
        restatements = pivot[(pivot["difference_rm"].abs() > 0.5)
                             & (pivot["original_sheet"] != pivot["restated_sheet"])].copy()
        restatements["pct_change"] = (
            restatements["difference_rm"] / restatements["original_rm"].abs().replace(0, pd.NA) * 100
        )

    df = (df.sort_values("priority", kind="stable")
            .drop_duplicates(key, keep="last")
            .drop(columns=["priority", "dedup_key"])
            .sort_values(["statement", "entity", "year", "label_key"])
            .reset_index(drop=True))
    return df, restatements


def validate(df: pd.DataFrame) -> pd.DataFrame:
    """Check that the parsed balance sheet still balances, per entity-year."""
    wide = (df[df["metric"].notna()]
            .pivot_table(index=["entity", "year"], columns="metric",
                         values="value_rm", aggfunc="first"))
    checks = []
    for (entity, year), row in wide.iterrows():
        assets = row.get("total_assets")
        liabs = row.get("total_liabilities")
        equity = row.get("total_equity")
        if pd.isna(assets) or pd.isna(liabs) or pd.isna(equity):
            continue
        diff = assets - (liabs + equity)
        checks.append({
            "entity": entity, "year": year,
            "total_assets": assets, "equity_plus_liabilities": liabs + equity,
            "difference_rm": diff, "balances": abs(diff) < 1.0,
        })
    return pd.DataFrame(checks).sort_values(["entity", "year"])


def main() -> None:
    ap = argparse.ArgumentParser(description="Flatten the SAA AFS workbook.")
    ap.add_argument("--source", default="SAA_Financial_statement.xlsx")
    ap.add_argument("--outdir", default="data")
    args = ap.parse_args()

    source = Path(args.source)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"Reading {source} ...")
    tidy, restatements = build_dataset(source)

    tidy_path = outdir / "saa_financials_tidy.csv"
    tidy.to_csv(tidy_path, index=False)
    print(f"\nWrote {tidy_path}  ({len(tidy)} rows, "
          f"{tidy['year'].min()}-{tidy['year'].max()})")

    mapped = tidy[tidy["metric"].notna()]
    wide = (mapped.pivot_table(index=["entity", "year"], columns="metric",
                               values="value_rm", aggfunc="first")
            .reset_index())
    wide_path = outdir / "saa_metrics_wide.csv"
    wide.to_csv(wide_path, index=False)
    print(f"Wrote {wide_path}  ({len(wide)} entity-years, "
          f"{len(wide.columns) - 2} metrics)")

    if not restatements.empty:
        r_path = outdir / "saa_restatements.csv"
        restatements.to_csv(r_path, index=False)
        print(f"Wrote {r_path}  ({len(restatements)} restated line items)")

    checks = validate(tidy)
    if not checks.empty:
        failed = checks[~checks["balances"]]
        print(f"\nBalance-sheet integrity: {len(checks) - len(failed)}/{len(checks)} balance")
        if not failed.empty:
            print(failed.to_string(index=False))


if __name__ == "__main__":
    main()
